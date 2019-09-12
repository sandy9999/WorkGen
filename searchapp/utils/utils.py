import openpyxl as op
from collections import defaultdict
from datetime import datetime
from random import sample


def default_to_regular(d):
    """
        This function converts a nested defaultdict to a dict

        :param d: defaultdict
        :type d: defaultdict
        :returns: equivalent dict
        :rtype: dict
    """
    if isinstance(d, defaultdict):
        d = {k: default_to_regular(v) for k, v in d.items()}
    return d


def convert_marker_data(marker_path, subject_breakup):
    """
        This function takes the path of the worksheet as an input along with the subject_breakup and returns
        a dictionary of names to score in each question type.

        Note: Please do not leave any mark field empty. If he/she got a zero, enter zero.

        :param marker_path: path of the excel file of the tracker
        :param subject_breakup: dictionay of QuestionType => [total_questions, no_to_attempt] for the given subject
        :type marker_path: str
        :type subject_breakup: str
        :returns: tuple of Dictionary of student_name => (Dictionary of question_type => [(marks_secured, question_number, chapter_number)])
                    along with chapter numbers asked in that test
        :rtype: dict

        :example subject_breakup:
            {
                '1A': [5, 5],
                '1B': [5, 5],
                '2': [7, 5],
                '3': [7, 5],
                '5': [2, 2]
            }
        :example of return:
            (
                {
                    'Akshay': {
                        '1A': [(1,1,1), (1,2,1), (1,3,1), (1,4,1), (1,5,1)],
                        '1B': ...
                        '2': ...
                        '3': ...
                        '5':  ...
                    }
                },
                [1,2,3,4,5]
            )
    """
    marker = op.load_workbook(marker_path).worksheets[0]
    row_no = -1
    student_to_answer = {}
    for row in marker.rows:
        row_no += 1

        # the first row is the header which has question types
        if row_no == 0:
            header_row = row
        # the second row has the chaper numbers
        elif row_no == 1:
            chapter_row = row
        # this is for every other row that contains data
        else:
            student = row[0].value
            if not student:
                break
            student_to_answer[student] = {}
            question_type_start_row = 1
            question_type = header_row[question_type_start_row].value
            while (question_type != None):
                #  Incase you get a string like '1A' as your question type, thats fine, else if you get a float like 2.0, convert to '2'
                question_type = question_type if type(
                    question_type) == str else str(int(question_type))
                total_questions, questions_to_attempt = subject_breakup[question_type]
                questions_of_given_type = []
                for i in range(question_type_start_row, question_type_start_row + total_questions):
                    question_number = i - question_type_start_row + 1
                    chapter_number = int(chapter_row[i].value)
                    marks_secured = row[i].value or 0.0
                    questions_of_given_type.append(
                        (marks_secured, question_number, chapter_number))
                questions_of_given_type.sort(key=lambda x: x[0], reverse=True)
                questions_of_given_type = questions_of_given_type[:questions_to_attempt]
                student_to_answer[student][question_type] = questions_of_given_type
                question_type_start_row += total_questions
                question_type = header_row[question_type_start_row].value
    return student_to_answer, list(set([int(val.value) for val in chapter_row if type(val.value) == type(1.0)]))


def get_allowed_questions(data, allowed_qtypes, allowed_chapters):
    """
        This function gets the parsed marker data as input and filters out chapters and question types according
        to the selection made

        :param data: Dictionary of student_name => (Dictionary of question_type => [(marks_secured, question_number, chapter_number)])
        :param allowed_qtypes: list of allowed question types
        :param allowed_chapters: list of allowed chapters
        :type data: dict
        :type allowed_qtypes: list
        :type allowed_chapters: list
        :returns: Updated data which has allowed_qtypes and allowed_chapters only
        :rtype: Dictionary of student_name => (Dictionary of question_type => [(marks_secured, question_number, chapter_number)])
    """
    updated_data = defaultdict(lambda: defaultdict(lambda: list))
    for student in data:
        qtypes = data[student]
        for qtype in qtypes:
            if qtype in allowed_qtypes:
                updated_data[student][qtype] = [
                    x for x in qtypes[qtype] if x[2] in allowed_chapters]
    updated_data = default_to_regular(updated_data)
    return updated_data


def get_customized_paper(marker_data):
    """
        This function returns a student wise split up of the customized paper given a particular marker data.
        Given the marks of a student, this function finds the ratio of marks_secured_for_a_question/total_marks_for_that_question
        We then take the three least scored chapters and generate questions based on those three chapters.

        :param marker_data: Dictionary of student_name => (Dictionary of question_type => [(marks_secured, question_number, chapter_number)])
        :type marker_data: dict
        :returns: dictionary of student names to chapter numbers they need more practice in
        :rtype: dict
    """
    student_to_chapter = {}
    for student in marker_data:
        student_data = marker_data[student]
        lowest_scored_chapters = []
        for question_type in student_data:
            results_per_questiontype = student_data[question_type]
            ratio_array = [(x[0] / float(question_type[0]), x[2]) for x in results_per_questiontype]
            for result in ratio_array:
                # if he/she scored less that 70% for that question, add that to the list of chapters to prepare
                if result[0] < 0.70:
                    lowest_scored_chapters.append(result)
        lowest_scored_chapters.sort(key=lambda x: x[0])
        student_to_chapter[student] = list(
            set([x[1] for x in lowest_scored_chapters]))
        student_to_chapter[student] = sample(
            student_to_chapter[student], min(3, len(student_to_chapter[student])))
    return student_to_chapter


def get_type_and_weightage(question_type):
    weightage = int(question_type[0])
    qtype = ord(question_type[1].lower()) - \
        ord('a') + 1 if len(question_type) == 2 else 1
    return weightage, qtype


def convert_question_bank(question_bank_path):
    """
        This function converts the excel file of the question bank to a mapping between chapter name to questions.

        :param question_bank_path: path of the question bank excel file
        :type question_bank_path: str
        :returns: dictionary of board =>( grade => (subject => (chapter_number => (dictionary of question_type => list of questions))))
        :rtype: dict

        :example return:
            {  'ICSE': {
                    10:{
                    'Science': {
                        (1, 'Gravitation') : {
                            '1A': [('Gravitation is caused by ______', 'Textbook'), ('There are ___ laws of Kepler', None), ('Earth has a radius of ___', None)],
                            ...
                            },
                            (2, 'Sun'): {
                            '1A': [],
                            '1B': [],
                            '2': [],
                            '3': [],
                            '4': []
                            }
                      }
                 }
            }
        }

    """
    question_bank = op.load_workbook(question_bank_path).worksheets[0]
    question_bank_dict = defaultdict(lambda: (defaultdict(lambda: defaultdict(
        lambda: defaultdict(lambda: defaultdict(lambda: list([])))))))
    row_no = -1
    for row in question_bank.rows:
        row_no += 1
        if row_no == 0:
            pass
        else:
            try:
                board = row[0].value
                grade = int(row[1].value)
                subject = row[2].value.lower()
                chapter_no = int(row[3].value)
                chapter_name = row[4].value
                question_type = row[5].value
                question_type = str(int(question_type)) if (type(question_type) == type(
                    1.0) or type(question_type) == type(1)) else question_type.lower()
                question_text = row[6].value
                question_source = row[7].value
                question_bank_dict[board][grade][subject][(
                    chapter_no, chapter_name)][question_type].append((question_text, question_source))
            except Exception as e:
                break
    return question_bank_dict


def generate_dummy_tracker(subject_name, subject_split, split_type_mapping):
    """
        Generate a workbook which has one tracket sheet with one dummy student
        :param subject_name: Name of subject to which subject_split belongs to
        :param subject_split: List of SubjectSplit objects
        :type subject_name: str
        :type subject_split: list

        :returns: Generated workbook object
        :rtype: openpyxl.Workbook
    """
    alphabets = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K",
                 "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]
    char_array = []
    for ch in alphabets:
        char_array += [ch + ele for ele in alphabets]
    char_array = alphabets + char_array

    dummy_workbook = op.Workbook()
    dummy_worksheet = dummy_workbook.active

    dummy_worksheet.title = "Tracker Record"
    dummy_worksheet["A1"] = subject_name
    dummy_worksheet["A2"] = "Chapter Number"
    dummy_worksheet["A3"] = "Akshay"

    bold_style = op.styles.Font(size=10, bold=True)
    dummy_worksheet["A1"].font = bold_style
    dummy_worksheet["A2"].font = bold_style

    thin = op.styles.Side(border_style="thin", color="000000")
    styles_border = op.styles.Border(
        top=thin, left=thin, right=thin, bottom=thin)
    styles_fillPeach = op.styles.PatternFill("solid", fgColor="fff2cc")
    styles_fillGreen = op.styles.PatternFill("solid", fgColor="e2efda")
    styles_centerAlignment = op.styles.Alignment(horizontal='center')

    style_counter = 0
    start_position = 1
    end_position = 1
    for split_object in subject_split:
        start_position = end_position
        end_position = start_position + split_object.total_questions

        style_counter += 1
        styles_currentFill = styles_fillPeach if style_counter % 2 else styles_fillGreen

        dummy_worksheet.merge_cells("{}1:{}1".format(char_array[start_position], char_array[end_position - 1]))
        dummy_worksheet["{}1".format(char_array[start_position])] = "{} Mark - {}".format(
            split_object.question_weightage,
            split_type_mapping[split_object.question_type - 1][1],
        )
        dummy_worksheet["{}1".format(char_array[start_position])].alignment = styles_centerAlignment
        dummy_worksheet["{}1".format(char_array[start_position])].border = styles_border
        dummy_worksheet["{}1".format(char_array[start_position])].fill = styles_currentFill

        for i in range(start_position, end_position):
            dummy_worksheet["{}2".format(char_array[i])] = 0
            dummy_worksheet["{}2".format(char_array[i])].border = styles_border
            dummy_worksheet["{}2".format(
                char_array[i])].fill = styles_currentFill

            dummy_worksheet["{}3".format(char_array[i])] = 0
            dummy_worksheet["{}3".format(char_array[i])].border = styles_border
            dummy_worksheet["{}3".format(
                char_array[i])].fill = styles_currentFill

    return dummy_workbook


if __name__ == "__main__":
    science_breakup = {
        '1A': [5, 5],
        '1B': [5, 5],
        '2': [7, 5],
        '3': [7, 5],
        '5': [2, 2]
    }
    data = convert_marker_data("data.xlsx", science_breakup)
    print(data)
